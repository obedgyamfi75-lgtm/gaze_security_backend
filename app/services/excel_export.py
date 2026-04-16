"""
GAZE Security Platform - Excel Export Service
One-way export of findings to Excel format
"""
import os
import tempfile
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Finding, Assessment, Severity, FindingStatus


# Severity colors
SEVERITY_COLORS = {
    'critical': 'FF0000',  # Red
    'high': 'FF6600',      # Orange
    'medium': 'FFCC00',    # Yellow
    'low': '00CC00',       # Green
    'informational': '0066FF',  # Blue
}

# Status colors
STATUS_COLORS = {
    'open': 'FF0000',
    'in_progress': 'FF9900',
    'remediated': '00CC00',
    'accepted': '0066FF',
    'false_positive': '999999',
    'duplicate': '999999',
}


class ExcelExportService:
    """Service for exporting findings to Excel"""
    
    @staticmethod
    def export_findings(assessment_id: Optional[str] = None) -> str:
        """
        Export findings to Excel file.
        
        Args:
            assessment_id: Optional assessment ID to filter by
        
        Returns:
            Path to generated Excel file
        """
        # Create workbook
        wb = Workbook()
        
        # Create findings sheet
        ws_findings = wb.active
        ws_findings.title = "Findings"
        
        # Query findings
        query = Finding.query
        if assessment_id:
            query = query.filter(Finding.assessment_id == assessment_id)
        
        findings = query.order_by(Finding.severity, Finding.created_at.desc()).all()
        
        # Write findings sheet
        ExcelExportService._write_findings_sheet(ws_findings, findings)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        ExcelExportService._write_summary_sheet(ws_summary, findings)
        
        # Create SLA tracking sheet
        ws_sla = wb.create_sheet("SLA Tracking")
        ExcelExportService._write_sla_sheet(ws_sla, findings)
        
        # Save to temp file
        fd, filepath = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(filepath)
        
        return filepath
    
    @staticmethod
    def _write_findings_sheet(ws, findings):
        """Write findings to worksheet"""
        # Headers
        headers = [
            'ID', 'Title', 'Severity', 'Status', 'CVSS', 'CWE', 'CVE',
            'OWASP', 'Affected Component', 'Affected URL', 'Description',
            'Impact', 'Recommendation', 'SLA Due Date', 'SLA Status',
            'Created', 'Remediated', 'Assessment', 'Asset'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Write data
        for row, finding in enumerate(findings, 2):
            ws.cell(row=row, column=1, value=str(finding.id)[:8])
            ws.cell(row=row, column=2, value=finding.title)
            
            # Severity with color
            sev_cell = ws.cell(row=row, column=3, value=finding.severity.value.upper())
            sev_cell.fill = PatternFill(
                start_color=SEVERITY_COLORS.get(finding.severity.value, 'FFFFFF'),
                end_color=SEVERITY_COLORS.get(finding.severity.value, 'FFFFFF'),
                fill_type='solid'
            )
            sev_cell.font = Font(bold=True, color='FFFFFF' if finding.severity.value in ['critical', 'high'] else '000000')
            
            # Status with color
            status_cell = ws.cell(row=row, column=4, value=finding.status.value.replace('_', ' ').title())
            status_cell.fill = PatternFill(
                start_color=STATUS_COLORS.get(finding.status.value, 'FFFFFF'),
                end_color=STATUS_COLORS.get(finding.status.value, 'FFFFFF'),
                fill_type='solid'
            )
            
            ws.cell(row=row, column=5, value=finding.cvss_score)
            ws.cell(row=row, column=6, value=finding.cwe_id)
            ws.cell(row=row, column=7, value=finding.cve_id)
            ws.cell(row=row, column=8, value=finding.owasp_category)
            ws.cell(row=row, column=9, value=finding.affected_component)
            ws.cell(row=row, column=10, value=finding.affected_url)
            ws.cell(row=row, column=11, value=finding.description[:500] if finding.description else '')
            ws.cell(row=row, column=12, value=finding.impact[:500] if finding.impact else '')
            ws.cell(row=row, column=13, value=finding.recommendation[:500] if finding.recommendation else '')
            ws.cell(row=row, column=14, value=finding.sla_due_date.isoformat() if finding.sla_due_date else '')
            
            # SLA status with conditional formatting
            sla_cell = ws.cell(row=row, column=15, value=finding.sla_status)
            if finding.is_overdue:
                sla_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sla_cell.font = Font(bold=True, color='FFFFFF')
            
            ws.cell(row=row, column=16, value=finding.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=17, value=finding.remediated_at.strftime('%Y-%m-%d') if finding.remediated_at else '')
            ws.cell(row=row, column=18, value=finding.assessment.name if finding.assessment else '')
            ws.cell(row=row, column=19, value=finding.assessment.asset.name if finding.assessment and finding.assessment.asset else '')
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Wider columns for text
        ws.column_dimensions['B'].width = 40  # Title
        ws.column_dimensions['K'].width = 50  # Description
        ws.column_dimensions['L'].width = 40  # Impact
        ws.column_dimensions['M'].width = 40  # Recommendation
    
    @staticmethod
    def _write_summary_sheet(ws, findings):
        """Write summary statistics"""
        ws.cell(row=1, column=1, value="Security Findings Summary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        
        # Severity breakdown
        ws.cell(row=4, column=1, value="Severity Breakdown")
        ws.cell(row=4, column=1).font = Font(bold=True)
        
        severity_counts = {}
        for s in Severity:
            severity_counts[s.value] = sum(1 for f in findings if f.severity == s)
        
        row = 5
        for sev, count in severity_counts.items():
            ws.cell(row=row, column=1, value=sev.title())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=SEVERITY_COLORS.get(sev, 'FFFFFF'),
                end_color=SEVERITY_COLORS.get(sev, 'FFFFFF'),
                fill_type='solid'
            )
            row += 1
        
        # Status breakdown
        ws.cell(row=row + 1, column=1, value="Status Breakdown")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        
        status_counts = {}
        for s in FindingStatus:
            status_counts[s.value] = sum(1 for f in findings if f.status == s)
        
        row += 2
        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=status.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=count)
            row += 1
    
    @staticmethod
    def _write_sla_sheet(ws, findings):
        """Write SLA tracking sheet"""
        headers = ['Title', 'Severity', 'Status', 'SLA Due Date', 'Days Remaining', 'SLA Status']
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Filter to open findings with SLA
        open_findings = [f for f in findings if f.status in [FindingStatus.OPEN, FindingStatus.IN_PROGRESS] and f.sla_due_date]
        open_findings.sort(key=lambda f: f.sla_due_date)
        
        for row, finding in enumerate(open_findings, 2):
            ws.cell(row=row, column=1, value=finding.title)
            ws.cell(row=row, column=2, value=finding.severity.value.upper())
            ws.cell(row=row, column=3, value=finding.status.value.replace('_', ' ').title())
            ws.cell(row=row, column=4, value=finding.sla_due_date.isoformat())
            ws.cell(row=row, column=5, value=finding.days_until_due)
            
            sla_cell = ws.cell(row=row, column=6, value=finding.sla_status)
            if finding.is_overdue:
                sla_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sla_cell.font = Font(bold=True, color='FFFFFF')
            elif finding.days_until_due and finding.days_until_due <= 7:
                sla_cell.fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
